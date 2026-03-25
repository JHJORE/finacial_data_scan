import openpyxl
from datetime import datetime
from pathlib import Path

from .models import Company


def _normalize_header_name(header: object) -> str:
    """Normalize Excel column header for matching.

    Excel sometimes embeds a carriage return in header cells; in the file this can
    appear as a literal ``_x000d_`` suffix (OOXML escape for CR) on the column name.
    """
    if header is None:
        return ""
    s = str(header).strip().lower()
    s = s.replace("_x000d_", "").replace("\r", "").replace("\n", "")
    return s


def _normalize_excel_cell_str(value: object) -> str:
    """Strip stray CR/LF and OOXML ``_x000d_`` artifacts from string cell values."""
    if value is None:
        return ""
    s = str(value).strip().replace("_x000d_", "").replace("\r", "").replace("\n", "")
    return s


def load_companies(filepath: Path) -> list[Company]:
    """Load companies from Excel file with 'acquirer', 'ticker', and 'first_entry' columns."""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    name_col = None
    ticker_col = None
    entry_col = None

    for i, header in enumerate(headers):
        key = _normalize_header_name(header)
        if key == "acquirer":
            name_col = i
        elif key == "ticker":
            ticker_col = i
        elif key == "first_entry":
            entry_col = i

    if name_col is None or ticker_col is None or entry_col is None:
        raise ValueError(
            f"Excel file must have 'acquirer', 'ticker', and 'first_entry' columns. Found: {headers}"
        )

    companies = []
    seen_slugs = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        name = _normalize_excel_cell_str(row[name_col])
        ticker = _normalize_excel_cell_str(row[ticker_col])
        raw_entry = row[entry_col]
        if isinstance(raw_entry, datetime):
            first_entry = raw_entry
        else:
            first_entry = _normalize_excel_cell_str(raw_entry)

        if not name or not ticker or not first_entry:
            continue

        company = Company.from_row(name, ticker, first_entry)

        # Deduplicate by slug (now includes year, so same company + different year = different slug)
        if company.slug not in seen_slugs:
            seen_slugs.add(company.slug)
            companies.append(company)

    wb.close()
    return companies
